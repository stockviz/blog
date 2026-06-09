#!/usr/bin/env python3
"""Extract Indian equity portfolio from PPFAS portfolio disclosure files.

Reads all .xls/.xlsx files from downloads/ and outputs a single CSV:
  date, name, isin, industry, pct_net_assets

Handles three format eras:
  1. 2013–2018: Single-sheet consolidated (only Flexi Cap existed)
  2. 2021-Sep – 2022-Oct: Multi-fund consolidated with Index sheet
  3. 2022-Nov – present: PPFCF-specific single-sheet
"""

import csv
import json
import os
import re
import sys
from datetime import datetime

import xlrd
import openpyxl


def load_config():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, "config", "config.json")
    with open(config_path) as f:
        cfg = json.load(f)
    cfg["_script_dir"] = script_dir
    cfg["download_dir"] = os.path.join(script_dir, cfg["download_dir"])
    return cfg


# ---- Section boundary markers ----
STOP_SECTIONS = [
    "foreign securities", "foreign equity", "overseas", "foreign etf",
    "arbitrage", "special situation",
    "debt instruments", "debt and money market", "money market",
    "cash", "cash and cash equivalent", "cash & cash equivalent",
    "total", "nil", "n i l",
    "derivative", "rights", "warrants",
    "reits", "reit", "invits",
    "preference", "mutual fund units",
    "grand total", "net current asset",
]

# Section headers that mark the START of equity section
EQUITY_SECTION_MARKERS = [
    "equity & equity related", "equity and equity related",
    "core equity", "equity",
    "a) listed", "(a) listed",
    "listed / awaiting listing",
    "listed/awaiting listing",
]

STOP_AFTER_EQUITY = [
    "b) unlisted", "(b) unlisted",
    "c) reits", "(c) reits",
    "d) foreign", "(d) foreign",
    "e)", "(e)",
]


def is_indian_isin(isin):
    """Indian equity ISINs start with INE (or INF for mutual funds, but we want INE)."""
    return bool(isin) and str(isin).strip().upper().startswith("INE")


def looks_like_stop(text):
    """Check if a row text looks like a section boundary we should stop at."""
    if not text:
        return False
    t = str(text).strip().lower()
    if not t:
        return False
    for marker in STOP_SECTIONS + STOP_AFTER_EQUITY:
        if t.startswith(marker) or t == marker:
            return True
    return False


def parse_date_from_filename(fname):
    """Extract a YYYY-MM-DD date from the filename."""
    fname_lower = fname.lower()

    # Pattern 1: "Month Day, Year" like "April_30_2026" or "June_30_2023"
    m = re.search(
        r'(january|february|march|april|may|june|july|august|september|october|november|december)[_\s-]+(\d{1,2})[_\s,-]+(20\d{2})',
        fname_lower,
    )
    if m:
        dt = datetime.strptime(f"{m.group(1)} {m.group(2)} {m.group(3)}", "%B %d %Y")
        return dt.strftime("%Y-%m-%d")

    # Pattern 2: abbreviated month like "Sept_30_2021" or "Oct_31_2021"
    m = re.search(
        r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[_\s-]+(\d{1,2})[_\s,-]+(20\d{2})',
        fname_lower,
    )
    if m:
        dt = datetime.strptime(f"{m.group(1)} {m.group(2)} {m.group(3)}", "%b %d %Y")
        return dt.strftime("%Y-%m-%d")

    # Pattern 3: "February-28-2015"
    m = re.search(
        r'(january|february|march|april|may|june|july|august|september|october|november|december)[_\s-]+(\d{1,2})[_\s-]+(20\d{2})',
        fname_lower,
    )
    if m:
        dt = datetime.strptime(f"{m.group(1)} {m.group(2)} {m.group(3)}", "%B %d %Y")
        return dt.strftime("%Y-%m-%d")

    # Pattern 4: Just month-year like "June-2013" (use last day of month)
    m = re.search(
        r'(january|february|march|april|may|june|july|august|september|october|november|december)[_\s-]+(20\d{2})',
        fname_lower,
    )
    if m:
        dt = datetime.strptime(f"{m.group(1)} {m.group(2)}", "%B %Y")
        # Last day of month
        if dt.month == 12:
            dt = dt.replace(day=31)
        else:
            dt = dt.replace(month=dt.month + 1, day=1)
            from datetime import timedelta
            dt = dt - timedelta(days=1)
        return dt.strftime("%Y-%m-%d")

    return None


def find_flexi_cap_sheet(wb, fname):
    """Find the sheet containing Flexi Cap Fund portfolio."""
    if fname.endswith(".xlsx"):
        sheet_names = wb.sheetnames
    else:
        sheet_names = wb.sheet_names()

    # PPFCF-specific files have a single "PPFCF" sheet
    if "PPFCF" in sheet_names:
        return "PPFCF"

    # Consolidated files have "Index" sheet with fund mappings
    if "Index" in sheet_names:
        if fname.endswith(".xlsx"):
            idx_ws = wb["Index"]
            for row in idx_ws.iter_rows(min_row=2, values_only=True):
                if row[2] and "flexi cap" in str(row[2]).lower():
                    return row[1]  # Short code like PPLTVF
        else:
            idx_ws = wb.sheet_by_name("Index")
            for r in range(1, idx_ws.nrows):
                name = str(idx_ws.cell_value(r, 2))
                if "flexi cap" in name.lower():
                    return str(idx_ws.cell_value(r, 1))

    # Old single-sheet files (2013–2018)
    if "PPLTVF" in sheet_names:
        return "PPLTVF"
    if "Portfolio" in sheet_names:
        return "Portfolio"
    if "Sheet1" in sheet_names:
        return "Sheet1"

    return sheet_names[0]


def get_cell_value(ws, row_idx, col_idx, is_xlsx):
    """Get a cell value, handling both xlrd and openpyxl."""
    if is_xlsx:
        val = ws.cell(row=row_idx + 1, column=col_idx + 1).value
        return val if val is not None else ""
    else:
        return ws.cell_value(row_idx, col_idx)


def get_nrows_ncols(ws, is_xlsx):
    if is_xlsx:
        return ws.max_row, ws.max_column
    else:
        return ws.nrows, ws.ncols


def extract_old_format(ws, is_xlsx, date_str):
    """Extract from old single-sheet format (2013–2018)."""
    rows = []
    nrows, ncols = get_nrows_ncols(ws, is_xlsx)

    # Find the header row — look for "Name of Instrument" or "ISIN"
    header_row = None
    for r in range(nrows):
        row_text = " ".join(str(get_cell_value(ws, r, c, is_xlsx)).lower() for c in range(min(ncols, 8)))
        if "name of instrument" in row_text or "name of the instrument" in row_text:
            header_row = r
            break

    if header_row is None:
        # Fallback: look for ISIN column header
        for r in range(nrows):
            for c in range(min(ncols, 8)):
                if str(get_cell_value(ws, r, c, is_xlsx)).strip().upper() == "ISIN":
                    header_row = r
                    break
            if header_row is not None:
                break

    if header_row is None:
        return rows

    # Determine column positions from header
    col_name = None
    col_isin = None
    col_industry = None
    col_pct = None

    for c in range(ncols):
        hdr = str(get_cell_value(ws, header_row, c, is_xlsx)).strip().lower()
        if "name of" in hdr and "instrument" in hdr:
            col_name = c
        elif hdr == "isin":
            col_isin = c
        elif "industry" in hdr:
            col_industry = c
        elif "% to nav" in hdr or "% to aum" in hdr or "% to net" in hdr:
            col_pct = c

    if col_name is None or col_isin is None:
        return rows

    # Find the equity section
    in_equity = False
    for r in range(header_row + 1, nrows):
        # Check row text for section markers
        row_texts = []
        for c in range(min(ncols, 5)):
            val = str(get_cell_value(ws, r, c, is_xlsx)).strip().lower()
            if val:
                row_texts.append(val)
        row_text = " ".join(row_texts)

        if not in_equity:
            for marker in EQUITY_SECTION_MARKERS:
                if marker in row_text:
                    in_equity = True
                    break
            continue

        # Check stop markers
        if looks_like_stop(row_text):
            break

        # Get values
        name = str(get_cell_value(ws, r, col_name, is_xlsx)).strip()
        isin = str(get_cell_value(ws, r, col_isin, is_xlsx)).strip()
        industry = str(get_cell_value(ws, r, col_industry, is_xlsx)).strip() if col_industry is not None else ""
        pct_raw = get_cell_value(ws, r, col_pct, is_xlsx) if col_pct is not None else ""

        if not is_indian_isin(isin):
            continue
        if not name or name.lower() in ("nil", "n i l", "total", ""):
            continue

        try:
            pct = float(pct_raw) if pct_raw != "" else 0.0
        except (ValueError, TypeError):
            pct = 0.0

        rows.append([date_str, name, isin, industry, round(pct, 4)])

    return rows


def extract_consolidated_format(ws, is_xlsx, date_str):
    """Extract from consolidated multi-fund format (2021–2022) — PPLTVF sheet."""
    rows = []
    nrows, ncols = get_nrows_ncols(ws, is_xlsx)

    # In this format:
    # Row 0-2: headers/fund info (first 2 cols are empty)
    # Row 3: column headers (col C=Name, D=ISIN, E=Industry, G=% to NAV/AUM)
    # The actual columns are:
    #   col 2 (C): Name of the Instrument / Issuer
    #   col 3 (D): ISIN
    #   col 4 (E): Rating / Industry ^
    #   col 6 (G): % to NAV (or % to AUM)

    col_name = 2
    col_isin = 3
    col_industry = 4
    col_pct = None

    # Find the actual % column from the header row
    for r in range(min(8, nrows)):
        for c in range(ncols):
            hdr = str(get_cell_value(ws, r, c, is_xlsx)).strip().lower()
            if "% to nav" in hdr or "% to aum" in hdr or "% to net" in hdr:
                col_pct = c
                break
        if col_pct is not None:
            break

    if col_pct is None:
        return rows  # Can't find % column

    in_equity = False
    for r in range(nrows):
        row_texts = []
        for c in range(min(ncols, 5)):
            val = str(get_cell_value(ws, r, c, is_xlsx)).strip().lower()
            if val:
                row_texts.append(val)
        row_text = " ".join(row_texts)

        if not in_equity:
            if "equity & equity related" in row_text or "equity and equity related" in row_text:
                in_equity = True
            continue

        # Check stop markers
        if looks_like_stop(row_text):
            break

        name = str(get_cell_value(ws, r, col_name, is_xlsx)).strip()
        isin = str(get_cell_value(ws, r, col_isin, is_xlsx)).strip()
        industry = str(get_cell_value(ws, r, col_industry, is_xlsx)).strip()
        pct_raw = get_cell_value(ws, r, col_pct, is_xlsx)

        if not is_indian_isin(isin):
            continue
        if not name or name.lower() in ("nil", "n i l", "total", ""):
            continue

        try:
            pct = float(pct_raw) if pct_raw != "" else 0.0
        except (ValueError, TypeError):
            pct = 0.0

        rows.append([date_str, name, isin, industry, round(pct, 4)])

    return rows


def extract_ppfcf_format(ws, is_xlsx, date_str):
    """Extract from PPFCF-specific single-sheet format (2022-Nov onward)."""
    rows = []
    nrows, ncols = get_nrows_ncols(ws, is_xlsx)

    # Column B: Name, C: ISIN, D: Industry, F: % to Net Assets
    col_name = 1   # B
    col_isin = 2   # C
    col_industry = 3  # D
    col_pct = 5    # F (sometimes 6 depending on version)

    # Find the actual % column header
    for r in range(min(5, nrows)):
        for c in range(ncols):
            hdr = str(get_cell_value(ws, r, c, is_xlsx)).strip().lower()
            if "% to net" in hdr or "% to nav" in hdr or "% to aum" in hdr:
                col_pct = c
                break

    in_equity = False
    for r in range(nrows):
        row_texts = []
        for c in range(min(ncols, 5)):
            val = str(get_cell_value(ws, r, c, is_xlsx)).strip().lower()
            if val:
                row_texts.append(val)
        row_text = " ".join(row_texts)

        if not in_equity:
            if "equity & equity related" in row_text or "equity and equity related" in row_text:
                in_equity = True
            continue

        # Check stop markers
        if looks_like_stop(row_text) or "foreign securities" in row_text:
            break

        name = str(get_cell_value(ws, r, col_name, is_xlsx)).strip()
        isin = str(get_cell_value(ws, r, col_isin, is_xlsx)).strip()
        industry = str(get_cell_value(ws, r, col_industry, is_xlsx)).strip()
        pct_raw = get_cell_value(ws, r, col_pct, is_xlsx)

        if not is_indian_isin(isin):
            continue
        if not name or name.lower() in ("nil", "n i l", "total", ""):
            continue

        try:
            pct = float(pct_raw) if pct_raw != "" else 0.0
        except (ValueError, TypeError):
            pct = 0.0

        rows.append([date_str, name, isin, industry, round(pct, 4)])

    return rows


def extract_from_file(filepath):
    """Extract Indian equity holdings from a single file. Returns list of rows."""
    fname = os.path.basename(filepath)
    date_str = parse_date_from_filename(fname)
    if date_str is None:
        print(f"  SKIP: could not parse date from {fname}", file=sys.stderr)
        return []

    is_xlsx = filepath.endswith(".xlsx")

    try:
        if is_xlsx:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        else:
            wb = xlrd.open_workbook(filepath)
    except Exception as e:
        print(f"  ERROR opening {fname}: {e}", file=sys.stderr)
        return []

    # Find the right sheet
    sheet_name = find_flexi_cap_sheet(wb, fname)
    if sheet_name is None:
        print(f"  SKIP {fname}: no Flexi Cap sheet found", file=sys.stderr)
        return []

    if is_xlsx:
        ws = wb[sheet_name]
    else:
        ws = wb.sheet_by_name(sheet_name)

    nrows, ncols = get_nrows_ncols(ws, is_xlsx)

    # Detect format
    if sheet_name == "PPFCF":
        rows = extract_ppfcf_format(ws, is_xlsx, date_str)
    elif sheet_name in ("PPLTVF",):
        rows = extract_consolidated_format(ws, is_xlsx, date_str)
    else:
        # Old format: Portfolio or Sheet1
        rows = extract_old_format(ws, is_xlsx, date_str)

    if is_xlsx:
        wb.close()

    return rows


def main():
    cfg = load_config()
    download_dir = cfg["download_dir"]

    files = sorted(
        f for f in os.listdir(download_dir)
        if f.endswith(".xls") or f.endswith(".xlsx")
    )

    print(f"Processing {len(files)} files...")

    all_rows = []
    files_with_data = 0
    files_no_data = 0

    for i, fname in enumerate(files, 1):
        filepath = os.path.join(download_dir, fname)
        rows = extract_from_file(filepath)
        if rows:
            all_rows.extend(rows)
            files_with_data += 1
            if i % 20 == 0 or i == len(files):
                print(f"  [{i}/{len(files)}] {files_with_data} files with data, {len(all_rows)} holdings so far...")
        else:
            files_no_data += 1

    # Sort by date, then name
    all_rows.sort(key=lambda r: (r[0], r[1]))

    # Normalize pct_net_assets: some eras store as percentage (0-100), others as decimal (0-1).
    # Detect per-date: if sum > 50, it's percentage format → divide by 100.
    from collections import defaultdict
    by_date = defaultdict(list)
    for r in all_rows:
        by_date[r[0]].append(float(r[4]))
    for r in all_rows:
        total = sum(by_date[r[0]])
        scale = 100.0 if total > 50 else 1.0
        r[4] = round(float(r[4]) / scale, 6)

    # Write CSV
    output_path = os.path.join(cfg["_script_dir"], "indian_equity_lo_portfolio.csv")
    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["date", "name", "isin", "industry", "pct_net_assets"])
        writer.writerows(all_rows)

    print(f"\nDone!")
    print(f"  Files processed: {len(files)}")
    print(f"  Files with data: {files_with_data}")
    print(f"  Files without data: {files_no_data}")
    print(f"  Total holdings: {len(all_rows)}")
    print(f"  Date range: {all_rows[0][0]} to {all_rows[-1][0]}")
    print(f"  Output: {output_path}")


if __name__ == "__main__":
    main()
